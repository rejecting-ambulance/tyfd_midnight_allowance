#網頁操作
from selenium import webdriver
from selenium.webdriver.support.ui import Select #選單
from selenium.webdriver.common.by import By #定位
from selenium.webdriver.support.ui import WebDriverWait #等待載入
from selenium.webdriver.support import expected_conditions as EC #等待載入
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.common.exceptions import TimeoutException, NoSuchElementException, ElementNotInteractableException, StaleElementReferenceException

#excel操作
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
#輔助
from datetime import datetime
from datetime import timedelta
import time
import sys
import os
import pandas as pd
import json

"""



"""


def add_zero(cc):   #1位補0
    if len(cc) < 2:
        cc = '0' + cc
    return cc


def alert_click(driver):    #警告處理
    try:
        time.sleep(0.2)
        alert = driver.switch_to.alert
        print(f'  {alert.text}')
        alert.accept()
    except:
        pass


def setup_chrome_driver():
    options = Options()
    options.add_argument('--headless=new')
    options.add_argument('--disable-gpu')  # 禁用 GPU
    options.add_argument('--disable-software-rasterizer')
    options.add_argument('--no-sandbox')
    options.add_argument('--log-level=3')  # 僅顯示致命錯誤，關掉 Info/Warning
    options.add_experimental_option('excludeSwitches', ['enable-logging'])  # 關掉 DevTools 日誌
    
    driver = webdriver.Chrome(service=Service(), options=options)
    return driver


def get_exe_dir():  #取得 `.exe` 真正所在的目錄
    if getattr(sys, 'frozen', False):  # PyInstaller 打包後
        return os.path.dirname(sys.executable)
    else:
        return os.path.dirname(os.path.abspath(__file__))


def load_accounts(config_path="config.json"):   #載入
    with open(config_path, "r", encoding="utf-8") as f: 
        return json.load(f)

def get_setting_date(config):
    full_date = f'{config['year']}年{config["month"]}月'

    return full_date  # 沒找到就原樣回傳

def get_full_unit_name(name, config):
    flat_dic = flatten_unit_dic(config['unit_dic'])
    return flat_dic.get(name, name)  # 沒找到就原樣回傳

def flatten_unit_dic(unit_dic):
    flat_dic = {}
    for major, data in unit_dic.items():
        # 中隊名稱 → 完整名稱
        flat_dic[major] = data['full_name']
        # 分隊名稱 → 中隊-分隊完整名稱
        for sub, sub_full in data.get('sub_units', {}).items():
            flat_dic[sub] = f"{data['full_name']}{sub_full}"
    return flat_dic

def dropdown_by_value(id , value , driver, wait):   #一些瀏覽器操作區塊
    dropdown1 = wait.until(EC.visibility_of_element_located((By.ID, id)))   
    dropdown_sheet1 = Select(dropdown1)
    dropdown_sheet1.select_by_value(value)

def dropdown_by_index(id , value , driver, wait):
    dropdown1 = wait.until(EC.visibility_of_element_located((By.ID, id)))   
    dropdown_sheet1 = Select(dropdown1)
    dropdown_sheet1.select_by_index(value)

def dropdown_by_text(id , value , driver, wait):
    dropdown1 = wait.until(EC.visibility_of_element_located((By.ID, id)))   
    dropdown_sheet1 = Select(dropdown1)
    dropdown_sheet1.select_by_visible_text(value)

def click_by_id(id , driver, wait):
    button1 = driver.find_element(By.ID, id)
    button1.click()

def click_by_name(name , driver, wait):
    button1 = driver.find_element(By.NAME, name)
    button1.click()

def click_by_xpath(xpath , driver, wait):
    button1 = driver.find_element(By.XPATH, xpath)
    button1.click()

def select_click_xpath(xpath_1, xpath_2, driver, wait, msg1 =  '', msg2 = ''):
    try:
        element = wait.until(EC.element_to_be_clickable((By.XPATH, xpath_1)))
        element.click()
        print(msg1)

    except Exception as e:
        try:
            element = wait.until(EC.element_to_be_clickable((By.XPATH, xpath_2)))
            element.click()
            print(msg2)

        except Exception as e:
            print(f"{e}：{xpath_1}、{xpath_2} 都找不到")
    


def str_line(show):     #大區段分隔線
    max_len = 50
    dash_len = int((max_len - len(show))/2)
    dash = ''
    for i in range(dash_len):
        dash = dash + '='
    show = dash + show + dash + '\n'
    
    return show


def remove_duplicates(data):       #移除重複列
    seen = set()
    result = []
    for row in data:
        row_tuple = tuple(str(x) for x in row)
        if row_tuple not in seen:
            seen.add(row_tuple)
            result.append(row)
    return result

def insert_type(arr, new_value):
    if len(arr) < 9:
        arr.append(new_value)
    else:
        # 第9位是合併欄
        existing_values = set(arr[8].split(',')) if arr[8] else set()
        if new_value and new_value not in existing_values:
            arr[8] = f"{arr[8]},{new_value}".strip(',')
    return arr


def format_excel(output_path):
    wb = load_workbook(output_path)
    checkmark = '☑'

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # 設定欄寬
        col_widths = [15, 8, 15, 15, 15, 8, 8, 8, 18, 18]
        for i, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # 插入一行做標題，移動原有標題到第2行，資料到第3行
        ws.insert_rows(1)
        total_cols = ws.max_column
        merge_range = f"A1:{get_column_letter(total_cols)}1"
        ws.merge_cells(merge_range)
        title = f'桃園市政府消防局{get_full_unit_name(sheet_name, load_accounts())}{get_setting_date(load_accounts())}深夜危勞性勤務津貼個人申請表'
        ws['A1'] = title

        # 設定標題字型、大小、置中
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        # 設定標題字型、大小、置中（標楷體）
        ws['A1'].font = Font(name='標楷體', size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30


        # 淺藍底色填充（標題行第2行）
        header_fill = PatternFill(fill_type='solid', fgColor='C7DDFF')

        # 資料區統一字型、大小、置中、自動換行，加框線
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=total_cols):
            for cell in row:
                cell.font = Font(name = '標楷體',size=12)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = thin_border
                if cell.row == 2:
                    cell.fill = header_fill

        # 填入「確認申請」欄打勾符號（假設倒數第3欄）
        confirm_col = total_cols - 2
        for row in range(3, ws.max_row + 1):
            ws.cell(row=row, column=confirm_col).value = checkmark

        # 同一人姓名，合併「申請人核章」「備註」欄
        name_col = 2  # 第3欄：姓名
        apply_col = total_cols - 1  # 倒數第2欄：申請人核章
        remark_col = total_cols     # 最後一欄：備註

        current_person = None
        merge_start = 3  # 從資料開始行
        for row in range(3, ws.max_row + 2):  # +2 為了最後一段也觸發
            name = ws.cell(row, name_col).value if row <= ws.max_row else None

            if name != current_person:
                if row - merge_start > 1:
                    apply_range = f"{get_column_letter(apply_col)}{merge_start}:{get_column_letter(apply_col)}{row - 1}"
                    remark_range = f"{get_column_letter(remark_col)}{merge_start}:{get_column_letter(remark_col)}{row - 1}"
                    ws.merge_cells(apply_range)
                    ws.merge_cells(remark_range)
                current_person = name
                merge_start = row

    wb.save(output_path)
    print(f"✅ Excel Sign sheet exported successfully to： {output_path}")

def comapre_times(driver, wait, data, unit):      #爬蟲裡面的比對時間區段

    dropdown_by_value('_selYEAR',data['year'], driver, wait)
    dropdown_by_value('_selMONTH',data['month'], driver, wait)
    click_by_id('_btnQuery', driver, wait)    #點選查詢

    # 重試機制：最多等三次
    retry_count = 0
    max_retries = 3

    while retry_count < max_retries:
        try:
            wait.until(EC.element_to_be_clickable((By.ID, '_btnQuery')))
            time.sleep(2)
            print(f"✅ 查詢成功 (第 {retry_count + 1} 次嘗試)\n")
            break  # 成功就跳出迴圈
        except (TimeoutException, NoSuchElementException, ElementNotInteractableException, StaleElementReferenceException) as e:
            retry_count += 1
            if retry_count < max_retries:
                print(f"⚠️ 查詢失敗，正在重試... (第 {retry_count} 次失敗，還剩 {max_retries - retry_count} 次機會)")
                #print(f"錯誤類型: {type(e).__name__}")
                print(f"錯誤類型: {type(e).__name__}")
                time.sleep(1)  # 等待1秒後重試
            else:
                print(f"❌ 查詢失敗，已達最大重試次數 ({max_retries} 次)")
                print(f"錯誤訊息: {e}")
                return [], []  # 返回空的 wrong_array 和 table_content

    # Collect all rows first
    table = driver.find_element(By.XPATH, '//*[@id="frm"]/table/tbody/tr[5]/td/table/tbody')
    rows = table.find_elements(By.TAG_NAME, 'tr')
    
    table_content = []   #爬到的內容
    wrong_array = []    #要被剃除的內容
    
    #系統爬內容，抓紅底
    for index, row in enumerate(rows):
                    
        cells = row.find_elements(By.TAG_NAME, 'td')
        if not cells:
            cells = row.find_elements(By.TAG_NAME, 'th')

        cell_values = [cell.text.strip() for cell in cells]
        cell_values.insert(0, unit)

        # Skip header or empty rows
        if '開始時間' in cell_values or len(cell_values) < 7:
            continue

        # check the background color of the third cell
        bg_color = row.value_of_css_property('background-color')
        if bg_color.startswith('rgba(255, 147, 147'):
            print(f"⚠️ Rows {index - 1} duplicate : {cell_values[2]}")
            wrong_type = "案號重複(Red)"
            cell_values = insert_type(cell_values, wrong_type)

        table_content.append(cell_values)

    if len(table_content) == 0:
        print('🆗 查無資料\n')
        return [], []

    # ======= 第一階段：檢查每筆記錄的時間長短 =======
    origin_person = ''
    person_number = 0
    
    for i in range(len(table_content)):
        current = table_content[i]
        person_current = current[2]
        
        # 檢查是否為新的人員
        if origin_person != person_current:
            origin_person = person_current
            person_number = i
            # 只在第一個人之後印空行分隔
            #if i > 0:
            #    print('')
        
        # 解析時間
        try:
            start_current = datetime.strptime(current[4], "%Y/%m/%d %H:%M")
            end_current = datetime.strptime(current[5], "%Y/%m/%d %H:%M")
        except ValueError as e:
            print(f"⚠️ 時間格式錯誤 Row {i + 1}: {current[4]} 或 {current[5]}")
            wrong_type = f"時間格式錯誤 (no.{i + 1 - person_number})"
            table_content[i] = insert_type(table_content[i], wrong_type)
            continue
        
        # 檢查時間區間長短
        current_difference = end_current - start_current
        
        if current_difference <= timedelta(minutes=5):
            print(f"🔺 Row {i + 1} too short : {person_current}-{i + 1 - person_number}")
            wrong_type = f"區間過短(<5分鐘) (no.{i + 1 - person_number})"
            table_content[i] = insert_type(table_content[i], wrong_type)
        elif current_difference >= timedelta(hours=6):
            print(f"🔺 Row {i + 1} too long : {person_current}-{i + 1 - person_number}")
            wrong_type = f"區間過長(>6小時) (no.{i + 1 - person_number})"
            table_content[i] = insert_type(table_content[i], wrong_type)
        else:
            # 如果時間長短正常，先標記為空字串（後續可能因重疊而更新）
            table_content[i] = insert_type(table_content[i], '')

    # ======= 第二階段：檢查同一人員的時間重疊 =======
    origin_person = ''
    person_number = 0
    
    for i in range(len(table_content)):
        current = table_content[i]
        person_current = current[2]
        
        # 檢查是否為新的人員（第二階段不印空行）
        if origin_person != person_current:
            origin_person = person_current
            person_number = i
        
        # 如果不是最後一筆，且下一筆是同一人，則檢查重疊
        if i < len(table_content) - 1:
            next_row = table_content[i + 1]
            person_next = next_row[2]
            
            # 只有當下一筆是同一人時才檢查重疊
            if person_current == person_next:
                try:
                    start_current = datetime.strptime(current[4], "%Y/%m/%d %H:%M")
                    end_current = datetime.strptime(current[5], "%Y/%m/%d %H:%M")
                    start_next = datetime.strptime(next_row[4], "%Y/%m/%d %H:%M")
                except ValueError:
                    print(f"⚠️ 時間格式錯誤，無法比較 Row {i + 1} 和 {i + 2}")
                    wrong_type = f"時間格式錯誤 (no.{i + 1 - person_number}.{i + 2 - person_number})"
                    # 更新錯誤類型（可能已有其他錯誤）
                    if len(table_content[i]) >= 9 and table_content[i][8]:
                        table_content[i][8] = f"{table_content[i][8]},{wrong_type}".strip(',')
                    else:
                        table_content[i] = insert_type(table_content[i], wrong_type)
                    if len(table_content[i + 1]) >= 9 and table_content[i + 1][8]:
                        table_content[i + 1][8] = f"{table_content[i + 1][8]},{wrong_type}".strip(',')
                    else:
                        table_content[i + 1] = insert_type(table_content[i + 1], wrong_type)
                    continue
                
                # 檢查時間重疊
                if start_next <= end_current:
                    print(f"⚠️ Rows {i + 1} and {i + 2} overlap : {person_current}-{i + 1 - person_number}.{i + 2 - person_number}")
                    wrong_type = f"時間重疊 (no.{i + 1 - person_number}.{i + 2 - person_number})"
                    
                    # 更新錯誤類型（可能已有其他錯誤）
                    if len(table_content[i]) >= 9 and table_content[i][8]:
                        table_content[i][8] = f"{table_content[i][8]},{wrong_type}".strip(',')
                    else:
                        table_content[i] = insert_type(table_content[i], wrong_type)
                    if len(table_content[i + 1]) >= 9 and table_content[i + 1][8]:
                        table_content[i + 1][8] = f"{table_content[i + 1][8]},{wrong_type}".strip(',')
                    else:
                        table_content[i + 1] = insert_type(table_content[i + 1], wrong_type)
                else:
                    None
                    #print(f"✅ Rows {i + 1} and {i + 2} correct : {person_current}-{i + 1 - person_number}.{i + 2 - person_number}")
        else:
            # 最後一筆資料，只要時間長短正常就標記為正確
            if len(table_content[i]) < 9 or not table_content[i][8]:
                None
                #print(f"✅ Row {i + 1} correct : {person_current}-{i + 1 - person_number}")

    # 收集有錯誤的資料
    for row in table_content:
        if len(row) >= 9 and row[8]:
            wrong_array.append(row)

    # 檢查是否有錯誤，如果沒有錯誤顯示全對訊息
    if len(wrong_array) == 0:
        print("🟢 所有資料都正確！沒有發現任何錯誤。")
    else:
        print(f"🆖 共發現 {len(wrong_array)} 筆有錯誤的資料")

    print('\n')
    return wrong_array, table_content

def bug(data):
    print('\nWellcome to the fucking far kingddom - Shrek\n')
    #開啟Chrome瀏覽器、勤務系統
    #driver = webdriver.Chrome()
    driver = setup_chrome_driver()
    wait = WebDriverWait(driver, 30)  # 最長等待 10 秒

    driver.get('https://dutymgt.tyfd.gov.tw/tyfd119/login119')

    #登入操作
    username = driver.find_element(By.ID,"_txtUsername")
    password = driver.find_element(By.ID,"_txtPassword")
    username.send_keys(data['username'])
    password.send_keys(data['password'])

    click_by_name('login', driver, wait)  #點選登入
    try:
        wait.until(EC.presence_of_element_located((By.NAME, 'ehrFrame')))
        frameM = driver.find_element(By.NAME, 'ehrFrame')
    except Exception as e:
        input(f'{e}帳密錯誤，請確認config.json')
        raise Exception

    print(str_line('登入成功'))
    
    #切換到選單Frame|#frameset是組合，不是Frame
    frameM = driver.find_element(By.NAME, 'ehrFrame')
    driver.switch_to.frame(frameM)
    frameL1 = driver.find_element(By.NAME, 'sidemenuFrame')
    driver.switch_to.frame(frameL1)
    frameL2 = driver.find_element(By.NAME, 'contentSidemenu')
    driver.switch_to.frame(frameL2)

    #click_by_name('nodeIcon17', driver, wait)   #轉換左方選單
    #click_by_xpath('//*[@id="item23"]/tbody/tr/td[2]/a/font', driver, wait)   #深夜危勞按鈕

    select_click_xpath('//*[@id="folder17"]/tbody/tr[1]/td/a[1]/img', '//*[@id="folder14"]/tbody/tr[1]/td/a[1]/img', driver, wait)  #相關業務
    select_click_xpath('//*[@id="item23"]/tbody/tr/td[2]/a/font', '//*[@id="item20"]/tbody/tr/td[2]/a/font', driver, wait, '沒記錯的話，上次見到你是一個月前呢，大隊承辦\n', '好久不見，分隊承辦！\n')  #深夜危勞性勤務津貼個人申請表


    #轉換右方主要內容
    driver.switch_to.parent_frame()
    driver.switch_to.parent_frame()
    frameR1 = driver.find_element(By.NAME, 'contentFrame')
    driver.switch_to.frame(frameR1)
    frameR2 = driver.find_element(By.NAME, 'mainFrame')
    driver.switch_to.frame(frameR2)
    
    #查詢月份
    try:
        title = wait.until(EC.visibility_of_element_located((By.CLASS_NAME, 'title')))
    except:
        raise TimeoutError('連線逾時，請關閉後重新操作')

    except_sheet = []
    sign_sheet = []
    unit_error_count = {}  # 用來統計每個單位的錯誤筆數

    dropdown_id = '_selDeptno'
    
    for i in range(len(Select(driver.find_element(By.ID, dropdown_id)).options)):
    #for i in range(2): #for test
        # REFRESH the dropdown each loop
        dropdown_element = wait.until(EC.presence_of_element_located((By.ID, dropdown_id)))
        dropdown = Select(dropdown_element)

        # REFRESH the option list each loop
        option = dropdown.options[i]
        value = option.get_attribute('value')
        text = option.text.strip()

        # Select the option
        dropdown.select_by_value(value)
        print(f"🔽 Selecting: {text} (value={value})")

        wrong, money_sheet = comapre_times(driver, wait, data, text)
        
        # 統計該單位的錯誤筆數
        unit_error_count[text] = len(wrong)
        
        for row in wrong:
            except_sheet.append(row)
        for row in money_sheet:
            sign_sheet.append(row)

    driver.close()
    driver.quit()
   
    # ========== 查詢結束統計 ==========
    print(str_line('查詢結束統計'))
    total_errors = 0
    for unit, count in unit_error_count.items():
        if count > 0:
            print(f"🔺 {unit}: {count} 筆錯誤")
            total_errors += count
        else:
            None
            #print(f"🟢 {unit}: 無錯誤")
    
    print(f"\n📊 總計: {total_errors} 筆錯誤")
    if total_errors == 0:
        print("🎉 恭喜！所有單位都沒有錯誤！")
    print("=" * 50)
    
    clean_sheet = remove_duplicates(except_sheet)
    
    # Create a new workbook and sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Exception"

    # 欄位名稱
    header = ["單位","日期", "姓名", "勤務項目", "開始時間", "結束時間", "深夜勤務時數", "金額", "錯誤種類"]
    ws.append(header)

    if clean_sheet:
        # Write data rows
        for row in clean_sheet:
            ws.append(row)
    else:
        ws['A2'] = 'All Carrot'

    # 設定欄寬
    column_widths = {
        'A': 12,  # 單位
        'B': 12,  # 日期
        'C': 10,  # 姓名
        'D': 20,  # 勤務項目
        'E': 18,  # 開始時間
        'F': 18,  # 結束時間
        'G': 15,  # 深夜勤務時數
        'H': 8,   # 金額
        'I': 25   # 錯誤種類
    }

    # 應用欄寬設定
    for column, width in column_widths.items():
        ws.column_dimensions[column].width = width

    # 或者使用另一種方式設定欄寬（根據內容自動調整）
    # for column_cells in ws.columns:
    #     length = max(len(str(cell.value)) for cell in column_cells if cell.value)
    #     ws.column_dimensions[column_cells[0].column_letter].width = length + 2

    # Save Excel file
    output_path = os.path.join(os.getcwd(), f"深夜食堂 - 修ㄟ味噌湯(修正).xlsx")
    output_path = f"深夜食堂 - 修ㄟ味噌湯(修正).xlsx"
    wb.save(output_path)

    
    print(f"✅ Excel Correct sheet exported successfully to： {output_path}")

    # 建立 DataFrame
    df = pd.DataFrame(sign_sheet, columns=header)

    # 依單位分組
    grouped = df.groupby('單位')

    # 建立 Excel 檔
    output_path2 = '深夜食堂 - 千層明太子(簽名).xlsx'
    with pd.ExcelWriter(output_path2, engine='openpyxl') as writer:
        for unit, group in grouped:
            # 去掉「單位」欄位
            group_no_unit = group.drop(columns=['單位', '錯誤種類'])
            # 加上三個新欄位，預設空值
            group_no_unit['確認申請'] = ''
            group_no_unit['申請人核章'] = ''
            group_no_unit['備註'] = ''
            # 寫入分頁
            group_no_unit.to_excel(writer, sheet_name=unit, index=False)

    # ⭐ 用 openpyxl 處理合併單元格和加標題
    format_excel(output_path2)

    os.startfile(output_path)
    os.startfile(output_path2)

    input('\n===醒醒，天亮囉，下班啦===\n')
    

################################################主程式################################################
if __name__ == '__main__':

    config = load_accounts()
    bug(config)